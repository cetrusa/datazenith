import os
import pandas as pd
import time
import gc
import logging
import uuid
from decimal import Decimal
import unicodedata
from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from scripts.conexion import Conexion as con
from scripts.config import ConfigBasic
from scripts.text_cleaner import TextCleaner
import ast
import psutil

logger = logging.getLogger(__name__)


class InterfaceContable:
    """
    Clase para manejar la generación de la Interface Contable, con soporte para tareas asíncronas y progreso.
    Esta versión escribe directamente a Excel en chunks, sin usar SQLite.
    """

    def __init__(
        self,
        database_name,
        IdtReporteIni,
        IdtReporteFin,
        user_id,
        reporte_id,
        progress_callback=None,  # Añadido callback
    ):
        self.database_name = database_name
        self.IdtReporteIni = IdtReporteIni
        self.IdtReporteFin = IdtReporteFin
        self.user_id = user_id
        self.reporte_id = reporte_id
        self.progress_callback = progress_callback
        self.start_time = time.time()
        self.config = {}
        self.engine_mysql = None
        self.file_path = None
        self.file_name = None
        self.total_records_processed = 0
        self.total_records_estimate = 0
        self._update_progress("Inicializando", 1)
        try:
            self._configurar_conexiones()
        except Exception as e:
            logger.error(
                f"Error crítico durante la inicialización de InterfaceContable: {e}",
                exc_info=True,
            )
            self._update_progress(f"Error inicialización: {e}", 100)
            raise

    def _update_progress(
        self, stage, progress_percent, current_rec=None, total_rec=None
    ):
        if self.progress_callback:
            try:
                total = (
                    total_rec if total_rec is not None else self.total_records_estimate
                )
                current = (
                    current_rec
                    if current_rec is not None
                    else self.total_records_processed
                )
                safe_progress = max(0, min(100, int(progress_percent)))
                self.progress_callback(stage, safe_progress, current, total)
                logger.debug(
                    f"Progreso: {stage} - {safe_progress}% ({current:,}/{total:,})"
                )
            except Exception as e:
                logger.warning(f"Error al llamar progress_callback: {e}", exc_info=True)
        else:
            logger.debug(f"Progreso (sin callback): {stage} - {progress_percent}%")

    def _configurar_conexiones(self):
        self._update_progress("Configurando conexiones", 2)
        try:
            config_basic = ConfigBasic(self.database_name, self.user_id)
            self.config = config_basic.config
            required_keys = [
                "nmUsrIn",
                "txPassIn",
                "hostServerIn",
                "portServerIn",
                "dbBi",
            ]
            if not all(self.config.get(key) for key in required_keys):
                raise ValueError(
                    "Configuración de conexión a MySQL/MariaDB incompleta."
                )
            self.engine_mysql = con.ConexionMariadb3(
                str(self.config["nmUsrIn"]),
                str(self.config["txPassIn"]),
                str(self.config["hostServerIn"]),
                int(self.config["portServerIn"]),
                str(self.config["dbBi"]),
            )
        except Exception as e:
            logger.error(f"Error al configurar conexiones: {e}", exc_info=True)
            raise

    def _generate_sqlout(self, hoja):
        sql = self.config.get("nmProcedureInterface")
        if not sql:
            raise ValueError("Esta empresa no maneja Interface")
        if self.config["dbBi"] == "powerbi_tym_eje":
            return text(
                f"CALL {sql}('{self.IdtReporteIni}','{self.IdtReporteFin}','','{str(hoja)}',0,0,0);"
            )
        return text(
            f"CALL {sql}('{self.IdtReporteIni}','{self.IdtReporteFin}','','{str(hoja)}');"
        )

    def _write_query_to_excel(self, query, hoja, writer, chunksize=10000):
        """
        Escritura robusta a Excel por chunks, compatible con cualquier consulta/procedimiento.
        """
        stage_name = (
            f"Extrayendo y escribiendo datos de MySQL para hoja {hoja} (SIN CHUNKS)"
        )
        self._update_progress(stage_name, 10)
        total_processed = 0
        start_extract_time = time.time()

        def _normalize_col_name(name: object) -> str:
            raw = "" if name is None else str(name)
            raw = unicodedata.normalize("NFKD", raw)
            raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
            return raw.upper().strip()

        def _clean_string_preserve_spaces(value: str) -> str:
            # Quitar caracteres de control sin colapsar espacios (necesario para REPEAT(' ', 42) en DIRECCIÓN)
            if value is None:
                return ""
            if not isinstance(value, str):
                value = str(value)
            value = "".join(ch for ch in value if ord(ch) >= 32 or ch in "\t\n\r")
            return unicodedata.normalize("NFKD", value)
        try:
            # Ejecutar el query (puede ser CALL o SELECT)
            with self.engine_mysql.connect() as conn:
                result = conn.execute(query)
                columns = list(result.keys())
                all_rows = [
                    dict(row._mapping) if hasattr(row, "_mapping") else dict(row)
                    for row in result
                ]
                df_all = pd.DataFrame(all_rows)
                if not df_all.empty:
                    # Limpieza: sanitizar strings para evitar caracteres ilegales en Excel.
                    object_columns = df_all.select_dtypes(include=["object"]).columns
                    for col_name in object_columns:
                        is_terceros = str(hoja).strip().upper() == "TERCEROS"
                        is_direccion = _normalize_col_name(col_name) == "DIRECCION"

                        if is_terceros and is_direccion:
                            # IMPORTANTE: NO colapsar espacios; Siigo pide barrio + 42 espacios + dirección DIAN
                            df_all[col_name] = df_all[col_name].apply(
                                lambda v: _clean_string_preserve_spaces(v)
                                if isinstance(v, str)
                                else v
                            )
                        else:
                            df_all[col_name] = df_all[col_name].apply(
                                lambda v: TextCleaner.clean_for_excel(v)
                                if isinstance(v, str)
                                else v
                            )

                    # Asegurar que los Decimals se conviertan a float para que Excel los trate como números.
                    for col_name in df_all.columns:
                        df_all[col_name] = df_all[col_name].apply(
                            lambda value: float(value)
                            if isinstance(value, Decimal)
                            else value
                        )

                    # Siigo: en la hoja TERCEROS algunas columnas deben ser numéricas para que el
                    # number_format (miles/decimales) se refleje.
                    if str(hoja).strip().upper() == "TERCEROS":
                        integer_cols = {
                            "A",
                            "B",
                            "C",
                            "N",
                            "O",
                            "Q",
                            "R",
                            "S",
                            "T",
                            "U",
                            "V",
                            "W",
                            "Y",
                            "Z",
                            "AA",
                            "AB",
                            "AQ",
                            "AR",
                            "AT",
                            "AU",
                            "AV",
                            "AW",
                            "AY",
                            "BA",
                            "BB",
                            "BC",
                            "BD",
                            "BE",
                            "BF",
                            "BI",
                            "BJ",
                            "BL",
                            "BM",
                            "BV",
                            "BW",
                        }
                        decimal_cols = {"AM", "AO", "AP", "AX"}

                        for col_letter in integer_cols.union(decimal_cols):
                            idx = column_index_from_string(col_letter) - 1
                            if idx < len(df_all.columns):
                                series = df_all.iloc[:, idx]
                                # Mantener vacíos como NaN para que Excel quede en blanco
                                series = series.replace("", pd.NA)
                                df_all.iloc[:, idx] = pd.to_numeric(series, errors="coerce")

                    df_all.to_excel(
                        writer,
                        sheet_name=hoja,
                        startrow=0,
                        index=False,
                        header=True,
                    )
                    total_processed = len(df_all)
                else:
                    empty_df = pd.DataFrame(columns=columns)
                    empty_df.to_excel(writer, sheet_name=hoja, index=False)
        except SQLAlchemyError as e:
            logger.error(
                f"Error de base de datos durante la extracción para {hoja}: {e}",
                exc_info=True,
            )
            self._update_progress(f"Error BD en {hoja}: {e}", 100)
            raise
        except Exception as e:
            logger.error(
                f"Error inesperado durante la extracción para {hoja}: {e}",
                exc_info=True,
            )
            self._update_progress(f"Error en {hoja}: {e}", 100)
            raise
        self.total_records_processed = total_processed
        self._update_progress(
            f"Datos extraídos y escritos a Excel para hoja {hoja}",
            80,
            total_processed,
        )

    def _aplicar_formato_siigo(self, workbook, sheet_name, start_row=1):
        """Despacha el formato de salida de acuerdo con la hoja solicitada."""
        empresa = self.config.get("name") or self.database_name
        if sheet_name.strip().upper() == "TERCEROS":
            self._formato_terceros(workbook, sheet_name, empresa)
        else:
            self._formato_movimiento(workbook, sheet_name, empresa)

    def _formato_terceros(self, workbook, sheet_name, empresa):
        ws = workbook[sheet_name]
        header_row = 5  # según plantilla Siigo
        last_col_letter = "BX"  # columna final del modelo de terceros

        # Insertar filas para desplazar encabezado y datos
        ws.insert_rows(1, header_row - 1)

        # Títulos y merges exactos
        ws["A1"] = empresa
        ws["A2"] = "MODELO TERCEROS"
        for r in [1, 2, 3, 4]:
            ws.merge_cells(f"A{r}:{last_col_letter}{r}")
            ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"A{r}"].font = Font(name="Calibri", size=12, bold=True)

        header_font = Font(name="Calibri", size=10, bold=True)
        body_font = Font(name="Calibri", size=10)
        header_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Anchos de columnas según plantilla (primeras 12 columnas)
        template_widths = {
            "A": 36.5703125,
            "B": 29.85546875,
            "C": 28.0,
            "D": 87.28515625,
            "E": 16.7109375,
            "F": 21.85546875,
            "G": 22.5703125,
            "H": 23.140625,
            "I": 22.140625,
            "J": 50.28515625,
            "K": 36.0,
            "L": 50.5703125,
            "M": 84.28515625,
            "N": 6.140625,
            "O": 9.28515625,
            "P": 9.140625,
            "Q": 15.42578125,
            "R": 13.85546875,
            "U": 22.0,
            "V": 5.42578125,
            "W": 19.85546875,
            "X": 6.85546875,
            "Y": 23.85546875,
            "Z": 23.7109375,
            "AA": 23.140625,
            "AB": 19.7109375,
            "AC": 81.28515625,
            "AD": 47.85546875,
            "AE": 74.0,
            "AF": 27.7109375,
            "AG": 40.28515625,
            "AH": 14.5703125,
            "AI": 19.5703125,
            "AJ": 22.140625,
            "AK": 39.0,
            "AL": 42.28515625,
            "AM": 48.140625,
            "AN": 44.28515625,
            "AO": 50.140625,
            "AP": 20.0,
            "AQ": 18.85546875,
            "AR": 18.28515625,
            "AS": 16.5703125,
            "AT": 24.140625,
            "AU": 35.42578125,
            "AV": 12.42578125,
            "AW": 12.7109375,
            "AX": 39.85546875,
            "AY": 20.28515625,
            "AZ": 22.7109375,
            "BA": 18.28515625,
            "BB": 18.140625,
            "BC": 28.85546875,
            "BD": 20.85546875,
            "BE": 19.28515625,
            "BF": 19.140625,
            "BG": 18.28515625,
            "BH": 22.28515625,
            "BI": 33.85546875,
            "BJ": 52.85546875,
            "BK": 59.42578125,
            "BL": 21.5703125,
            "BM": 24.7109375,
            "BN": 41.5703125,
            "BO": 62.140625,
            "BP": 64.28515625,
            "BQ": 54.7109375,
            "BR": 84.5703125,
            "BS": 18.140625,
            "BT": 29.140625,
            "BU": 17.0,
            "BV": 16.85546875,
            "BW": 16.140625,
            "BX": 11.42578125,
            "BY": 11.42578125,
        }

        for col_idx, cell in enumerate(ws[header_row], start=1):
            col_letter = get_column_letter(col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_letter in template_widths:
                ws.column_dimensions[col_letter].width = template_widths[col_letter]
            else:
                ws.column_dimensions[col_letter].width = 22

        integer_format = r"##,##0_);[Red]\(##,##0\)"
        decimal_format = r"##,##0.00_);[Red]\(##,##0.00\)"
        integer_cols = {
            "A",
            "B",
            "C",
            "N",
            "O",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
            "Y",
            "Z",
            "AA",
            "AB",
            "AQ",
            "AR",
            "AT",
            "AU",
            "AV",
            "AW",
            "AY",
            "BA",
            "BB",
            "BC",
            "BD",
            "BE",
            "BF",
            "BI",
            "BJ",
            "BL",
            "BM",
            "BV",
            "BW",
        }
        decimal_cols = {"AM", "AO", "AP", "AX"}

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                col = cell.column_letter
                if col in decimal_cols:
                    cell.number_format = decimal_format
                elif col in integer_cols:
                    cell.number_format = integer_format
                else:
                    cell.number_format = "General"

        ws.freeze_panes = f"A{header_row + 1}"

    def _formato_movimiento(self, workbook, sheet_name, empresa):
        ws = workbook[sheet_name]
        header_row = 5  # según plantilla Siigo
        last_col_letter = "DS"  # columna final del modelo de movimiento contable

        ws.insert_rows(1, header_row - 1)

        ws["A1"] = empresa
        ws["A2"] = "MODELO PARA LA IMPORTACION DE MOVIMIENTO CONTABLE - MODELO GENERAL"
        for r in [1, 2, 3, 4]:
            ws.merge_cells(f"A{r}:{last_col_letter}{r}")
            ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"A{r}"].font = Font(name="Calibri", size=12, bold=True)

        header_font = Font(name="Calibri", size=10, bold=True)
        body_font = Font(name="Calibri", size=10)
        header_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        template_widths = {
            "A": 43.28515625,
            "B": 44.0,
            "C": 27.28515625,
            "D": 39.5703125,
            "E": 38.28515625,
            "F": 45.85546875,
            "G": 24.0,
            "H": 23.85546875,
            "I": 23.28515625,
            "J": 26.0,
            "K": 24.85546875,
            "L": 22.5703125,
            "M": 13.140625,
            "N": 20.42578125,
            "O": 24.7109375,
            "P": 4.7109375,
            "Q": 12.0,
            "R": 34.85546875,
            "S": 22.7109375,
            "T": 27.7109375,
            "U": 41.0,
            "V": 18.28515625,
            "W": 41.7109375,
            "Y": 47.0,
            "AB": 50.85546875,
            "AC": 63.85546875,
            "AD": 33.5703125,
            "AE": 40.0,
            "AF": 31.0,
            "AG": 31.140625,
            "AH": 38.0,
            "AI": 37.85546875,
            "AJ": 37.28515625,
            "AK": 29.28515625,
            "AL": 44.5703125,
            "AM": 46.7109375,
            "AN": 32.42578125,
            "AO": 32.5703125,
            "AP": 39.28515625,
            "AQ": 39.140625,
            "AR": 38.5703125,
            "AS": 32.42578125,
            "AT": 32.5703125,
            "AU": 39.28515625,
            "AV": 39.140625,
            "AW": 38.5703125,
            "AX": 32.42578125,
            "AY": 32.5703125,
            "AZ": 39.28515625,
            "BA": 39.140625,
            "BB": 38.5703125,
            "BC": 32.42578125,
            "BD": 32.5703125,
            "BE": 39.28515625,
            "BF": 39.140625,
            "BG": 38.5703125,
            "BH": 48.7109375,
            "BI": 42.28515625,
            "BJ": 32.85546875,
            "BK": 29.5703125,
            "BL": 29.42578125,
            "BM": 28.85546875,
            "BN": 42.85546875,
            "BO": 35.140625,
            "BP": 22.5703125,
            "BQ": 51.85546875,
            "BR": 36.28515625,
            "BS": 18.85546875,
            "BT": 15.28515625,
            "BU": 51.42578125,
            "BV": 19.28515625,
            "BW": 20.140625,
            "BX": 21.5703125,
            "BY": 11.85546875,
            "BZ": 16.85546875,
            "CA": 25.42578125,
            "CB": 28.85546875,
            "CC": 42.140625,
            "CD": 42.7109375,
            "CE": 39.42578125,
            "CF": 18.140625,
            "CH": 21.5703125,
            "CI": 63.5703125,
            "CJ": 27.5703125,
            "CK": 46.28515625,
            "CL": 46.140625,
            "CM": 37.42578125,
            "CN": 37.28515625,
            "CO": 36.5703125,
            "CP": 32.0,
            "CQ": 38.140625,
            "CR": 38.5703125,
            "CS": 25.140625,
            "CT": 34.85546875,
            "CU": 32.28515625,
            "CV": 34.85546875,
            "CW": 28.7109375,
            "CX": 45.85546875,
            "CY": 45.7109375,
            "CZ": 45.140625,
            "DA": 50.85546875,
            "DB": 34.85546875,
            "DC": 23.0,
            "DD": 13.42578125,
            "DE": 31.140625,
            "DF": 25.42578125,
            "DG": 18.0,
            "DH": 21.140625,
            "DI": 15.7109375,
            "DJ": 19.0,
            "DK": 12.7109375,
            "DL": 14.140625,
            "DM": 27.7109375,
            "DN": 29.140625,
            "DO": 40.7109375,
            "DP": 45.28515625,
            "DQ": 30.28515625,
            "DR": 24.0,
            "DS": 52.85546875,
            "DT": 11.42578125,
        }

        for col_idx, cell in enumerate(ws[header_row], start=1):
            col_letter = get_column_letter(col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_letter in template_widths:
                ws.column_dimensions[col_letter].width = template_widths[col_letter]
            else:
                ws.column_dimensions[col_letter].width = 20

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.font = body_font
                cell.border = thin_border
                cell.number_format = "General"
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.freeze_panes = f"A{header_row + 1}"

    def run(self):
        logger.info(
            "[InterfaceContable] INICIO del proceso de generación de interface contable (directo a Excel)"
        )
        try:
            logger.info(f"[InterfaceContable] Configuración cargada: {self.config}")
            txProcedureInterface_str = self.config.get("txProcedureInterface")
            logger.info(
                f"[InterfaceContable] txProcedureInterface_str: {txProcedureInterface_str}"
            )
            if isinstance(txProcedureInterface_str, str):
                try:
                    self.config["txProcedureInterface"] = ast.literal_eval(
                        txProcedureInterface_str
                    )
                except ValueError as e:
                    logger.error(
                        f"[InterfaceContable] Error al convertir txProcedureInterface a lista: {e}"
                    )
                    self.config["txProcedureInterface"] = []
            logger.info(
                f"[InterfaceContable] txProcedureInterface (list): {self.config['txProcedureInterface']}"
            )
            if not self.config["txProcedureInterface"] or not any(
                str(x).strip() for x in self.config["txProcedureInterface"]
            ):
                logger.warning("[InterfaceContable] No hay datos para procesar")
                return {"success": False, "error_message": "No hay datos para procesar"}

            # Generar nombre de archivo de salida
            ext = ".xlsx"
            empresa_nombre = self.config.get("name") or self.database_name
            empresa_slug = "".join(
                c if c.isalnum() or c in "-_" else "_" for c in str(empresa_nombre)
            )
            reporte_id_str = (
                f"_reporte_{self.reporte_id}"
                if hasattr(self, "reporte_id") and self.reporte_id
                else ""
            )
            user_id_str = f"_user_{self.user_id}" if self.user_id else ""
            self.file_name = f"Interface_Contable_{empresa_slug}_de_{self.IdtReporteIni}_a_{self.IdtReporteFin}{user_id_str}{reporte_id_str}{ext}"
            self.file_path = os.path.join("media", self.file_name)
            output_dir = os.path.dirname(self.file_path)
            os.makedirs(output_dir, exist_ok=True)

            # Usar openpyxl como engine para ExcelWriter
            with pd.ExcelWriter(self.file_path, engine="openpyxl") as writer:
                total_global_records = 0
                for idx, hoja in enumerate(
                    self.config["txProcedureInterface"], start=1
                ):
                    hoja = str(hoja).strip()
                    if not hoja:
                        continue
                    logger.info(
                        f"[InterfaceContable] Ejecutando query para extraer datos de la hoja: {hoja}"
                    )
                    try:
                        query = self._generate_sqlout(hoja)
                        logger.info(f"[InterfaceContable] Query generado: {query}")
                        self._write_query_to_excel(query, hoja, writer)
                        self._aplicar_formato_siigo(
                            workbook=writer.book, sheet_name=hoja, start_row=1
                        )
                        total_global_records += self.total_records_processed
                    except Exception as e:
                        logger.error(
                            f"[InterfaceContable] Error al procesar hoja {hoja}: {str(e)}",
                            exc_info=True,
                        )
                        continue
                    # Progreso global por hoja
                    if self.progress_callback:
                        percent = int(
                            (idx / len(self.config["txProcedureInterface"])) * 100
                        )
                        self.progress_callback(
                            f"Progreso global: {idx}/{len(self.config['txProcedureInterface'])} hojas",
                            percent,
                            hoja_idx=idx,
                            total_hojas=len(self.config["txProcedureInterface"]),
                        )

            execution_time = time.time() - self.start_time
            if total_global_records == 0:
                logger.warning(
                    "[InterfaceContable] No hay datos para mostrar en ninguna hoja."
                )
                return {
                    "success": False,
                    "error_message": "No hay datos para mostrar en ninguna hoja.",
                    "file_path": None,
                    "file_name": None,
                    "execution_time": execution_time,
                    "metadata": {"total_records": 0},
                }
            logger.info(f"[InterfaceContable] Archivo generado en: {self.file_path}")
            logger.info(
                f"[InterfaceContable] Proceso completado correctamente en {execution_time:.2f} segundos."
            )
            self._update_progress("Completado", 100, total_global_records)
            logger.info(
                f"InterfaceContable completada en {execution_time:.2f} segundos. Archivo: {self.file_path}"
            )
            return {
                "success": True,
                "message": f"Interface contable generada exitosamente en {execution_time:.2f} segundos.",
                "file_path": self.file_path,
                "file_name": self.file_name,
                "execution_time": execution_time,
                "metadata": {
                    "total_records": total_global_records,
                },
            }
        except Exception as e:
            execution_time = time.time() - self.start_time
            error_msg = (
                f"Error fatal en InterfaceContable.run: {type(e).__name__} - {e}"
            )
            logger.error(f"[InterfaceContable] ERROR: {error_msg}", exc_info=True)
            self._update_progress(f"Error fatal: {e}", 100)
            return {
                "success": False,
                "message": f"Error al generar la interface: {error_msg}",
                "error": error_msg,
                "file_path": None,
                "file_name": None,
                "execution_time": execution_time,
                "metadata": {},
            }
