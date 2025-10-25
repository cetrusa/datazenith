# -*- coding: utf-8 -*-
"""
Script de cargue de información de ventas a base de datos.
Soporta emojis y caracteres UTF-8 completos.
"""

import sys
import io

# Garantizar UTF-8 en stdout y stderr
if sys.platform == 'win32':
    # Windows: usar UTF-8 en lugar de cp1252
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 📦 CARGUE_INFOVENTAS_MAIN.PY
# ------------------------------------------------------------
# Descripción:
#   Script autónomo para cargar archivos de información de ventas
#   (.xlsx o .csv) hacia la base de datos correspondiente, sin
#   depender de Django. Utiliza ConfigBasic para resolver
#   automáticamente las credenciales y parámetros de conexión.
#
# ------------------------------------------------------------
# 🧰 Requisitos previos:
#   - Estructura de directorios:
#       scripts/
#           ├── conexion.py
#           ├── config.py
#           ├── config_repository.py
#           ├── cargue_infoventas_insert.py
#   - Configuración de acceso a la base:
#       Las credenciales deben estar registradas en las tablas
#       `powerbi_adm.conf_server` y `powerbi_adm.conf_tipo`,
#       o bien en `secret.json` si ConfigBasic lo permite.
#
# ------------------------------------------------------------
# ⚙️ Modo de uso (línea de comandos):
#
#   🔸 1. Procesar todos los archivos de una carpeta:
#       python cargue_infoventas_main.py --base bi_distrijass --carpeta "D:\Python\DataZenithBi\Info proveedores 2024"
#
#   🔸 2. Procesar un archivo único:
#       python cargue_infoventas_main.py --base bi_distrijass --archivo "D:\Python\DataZenithBi\Info proveedores 2024\infoventas_2025_01.xlsx"
#
#   🔸 3. Mostrar ayuda:
#       python cargue_infoventas_main.py --help
#
# ------------------------------------------------------------
# 🔍 Descripción de parámetros:
#
#   --base      Nombre lógico del entorno o conexión configurada.
#               (Ejemplo: bi_distrijass)
#
#   --carpeta   Ruta completa a una carpeta que contiene archivos
#               .xlsx o .csv a procesar. Los archivos se ejecutan
#               en orden alfabético.
#
#   --archivo   Ruta completa de un solo archivo a procesar.
#               Si se indica este parámetro, ignora --carpeta.
#
# ------------------------------------------------------------
# 🧠 Funcionamiento interno:
#
#   1. ConfigBasic obtiene las credenciales de conexión (usuario,
#      contraseña, host, puerto, base de datos) desde la fuente
#      configurada.
#
#   2. Se crea la conexión mediante ConexionMariadb3.
#
#   3. Se inicializa ConfigRepository y CargueInfoVentasInsert.
#
#   4. Cada archivo se procesa y se insertan los registros en la
#      tabla staging (infoventas), siguiendo la lógica interna
#      del cargador.
#
#   5. Se registra en consola el resumen de filas insertadas,
#      duplicadas y duración total del proceso.
#
# ------------------------------------------------------------
# 🧾 Ejemplo de salida esperada:
#
#   2025-10-03 14:10:25 [INFO] ⚙️ Configurando conexión para entorno 'bi_distrijass'...
#   2025-10-03 14:10:26 [INFO] ✅ Conectado a bi_distrijass en 181.49.241.226:3306
#   2025-10-03 14:10:26 [INFO] 🚀 Iniciando cargue del archivo: infoventas_2025_01.xlsx
#   2025-10-03 14:10:55 [INFO] ✅ Cargue completado con éxito.
#   2025-10-03 14:10:55 [INFO] 📊 Filas insertadas: 256334
#   2025-10-03 14:10:55 [INFO] 📦 Filas duplicadas: 112
#   2025-10-03 14:10:55 [INFO] 🕒 Duración total: 29.14 segundos
#
# ------------------------------------------------------------
# 🧩 Recomendaciones:
#
#   - Usa rutas absolutas (no relativas) para carpetas o archivos.
#   - Verifica antes del cargue que los nombres de columnas del
#     archivo coincidan con los esperados por CargueInfoVentasInsert.
#   - Puedes usar `time.sleep()` entre archivos para evitar sobrecarga
#     en el servidor si cargas grandes volúmenes.
#   - Ideal para programar en tareas automáticas (Windows Task
#     Scheduler o cron en Linux).
#
# ------------------------------------------------------------
# 👨‍💻 Autor: [Tu nombre o equipo]
# 📅 Última actualización: 2025-10-03
# ============================================================
# ============================================================
# 📦 CARGUE_INFOVENTAS_MAIN.PY (Versión final con mantenimiento automático)
# ============================================================

import os
import sys
import time
import logging
import argparse
from datetime import datetime
from scripts.cargue.cargue_infoventas_insert import CargueInfoVentasInsert
from sqlalchemy import text
from sqlalchemy.exc import OperationalError as SAOperationalError
from pymysql.err import OperationalError as PyMySQLOperationalError, InterfaceError as PyMySQLInterfaceError

# ============================================================
# 🎨 COLORES PARA TERMINAL (ANSI CODES)
# ============================================================
class TerminalColors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)

# ------------------------------------------------------------
# 🔍 Detectar fechas en el nombre del archivo
# ------------------------------------------------------------
def detectar_fechas_desde_nombre(nombre_archivo: str, archivo_path: str = None):
    """
    Extrae año y mes desde el nombre del archivo (ej: 2025-08 o 202508).
    Si no encuentra en el nombre, intenta extraer del Excel.
    """
    import re
    from calendar import monthrange
    
    # Intento 1: Buscar en el nombre del archivo
    match = re.search(r"(\d{4})[-_]?(\d{2})", nombre_archivo)
    if match:
        anio, mes = match.groups()
        try:
            fecha_ini = datetime(int(anio), int(mes), 1).date()
            fecha_fin = datetime(int(anio), int(mes), monthrange(int(anio), int(mes))[1]).date()
            return fecha_ini, fecha_fin
        except ValueError:
            pass
    
    # Intento 2: Si se proporciona ruta del archivo, buscar en metadatos del Excel
    if archivo_path and archivo_path.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(archivo_path, data_only=True)
            # Buscar en la primera hoja
            ws = wb.active
            # Buscar fechas en las primeras 10 filas y columnas
            for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10, values_only=True):
                for cell in row:
                    if cell:
                        cell_str = str(cell).strip()
                        # Buscar patrones de fecha
                        match = re.search(r"(\d{4})[-_/.](\d{2})", cell_str)
                        if match:
                            anio, mes = match.groups()
                            try:
                                fecha_ini = datetime(int(anio), int(mes), 1).date()
                                fecha_fin = datetime(int(anio), int(mes), monthrange(int(anio), int(mes))[1]).date()
                                logging.info(f"✅ Fechas detectadas desde Excel: {fecha_ini} → {fecha_fin}")
                                return fecha_ini, fecha_fin
                            except ValueError:
                                pass
        except Exception as e:
            logging.debug(f"No se pudo leer Excel para detectar fechas: {e}")
    
    return None, None


# ------------------------------------------------------------
# 🔁 Helper para ejecutar procedimientos con reintentos
# ------------------------------------------------------------
def ejecutar_procedimiento_con_reintentos(cargador, sentencia_sql: str, intentos: int = 3, espera_segundos: int = 45):
    """Ejecuta un procedimiento almacenado con reintentos y ajustes de timeout."""
    ultimo_error = None
    print(f"♻️ Preparando ejecución con hasta {intentos} intentos... [DEBUG]")
    logging.info(f"♻️ Preparando ejecución del procedimiento con hasta {intentos} intentos...")

    for intento in range(1, intentos + 1):
        print(f"   ▶️ Intento {intento}/{intentos}... [DEBUG]")
        logging.info(f"   ▶️ Intento {intento}/{intentos} de ejecución del procedimiento...")

        try:
            conn = cargador.engine_mysql_bi.raw_connection()
            try:
                conn.autocommit(True)
                cursor = conn.cursor()
                try:
                    ajustes_timeout = [
                        "SET SESSION wait_timeout = 7200",
                        "SET SESSION interactive_timeout = 7200",
                        "SET SESSION net_read_timeout = 600",
                        "SET SESSION net_write_timeout = 600",
                        "SET SESSION innodb_lock_wait_timeout = 900",
                    ]
                    for comando in ajustes_timeout:
                        cursor.execute(comando)

                    cursor.execute(sentencia_sql)

                    while True:
                        try:
                            filas = cursor.fetchall()
                            if filas:
                                print(f"📋 Resultados parciales: {filas} [DEBUG]")
                                logging.info(f"📋 Resultados parciales del procedimiento: {filas}")
                        except Exception:
                            pass

                        try:
                            tiene_mas = cursor.nextset()
                        except Exception:
                            tiene_mas = False

                        if not tiene_mas:
                            break

                    # Intentar commit, pero no fallar si no funciona
                    # (el procedimiento ya se ejecutó)
                    try:
                        conn.commit()
                    except Exception as commit_err:
                        logging.warning(f"   ⚠️ Aviso en commit: {commit_err} (procedimiento probablemente completado)")
                    
                    print(f"   ✅ Procedimiento finalizado en intento {intento} [DEBUG]")
                    logging.info(f"   ✅ Procedimiento finalizado en intento {intento}")
                    return True, None

                finally:
                    try:
                        cursor.close()
                    except Exception:
                        pass
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        except (PyMySQLOperationalError, PyMySQLInterfaceError, SAOperationalError) as db_err:
            # Normalizar código de error
            if isinstance(db_err, SAOperationalError) and hasattr(db_err, "orig"):
                codigo_error = getattr(db_err.orig, "args", [None])[0] if db_err.orig else None
                mensaje_error = str(db_err.orig)
            else:
                codigo_error = getattr(db_err, "args", [None])[0]
                mensaje_error = str(db_err)

            ultimo_error = db_err
            print(f"   ⚠️ Error de base de datos (código {codigo_error}): {mensaje_error} [DEBUG]")
            logging.warning(f"   ⚠️ Error de base de datos (código {codigo_error}): {mensaje_error}")

            if codigo_error == 0:
                print("   ℹ️ Código 0 recibido; se asume ejecución finalizada por cierre de resultados. [DEBUG]")
                logging.info("   ℹ️ Código 0 recibido; se asume ejecución finalizada por cierre de resultados.")
                return True, None

            if codigo_error in (2006, 2013, 1205) and intento < intentos:
                print(f"   ⏳ Reintentando en {espera_segundos} segundos... [DEBUG]")
                logging.info(f"   ⏳ Reintentando en {espera_segundos} segundos...")
                time.sleep(espera_segundos)
                continue
            else:
                break

        except Exception as error_general:
            ultimo_error = error_general
            print(f"   ❌ Error inesperado en intento {intento}: {error_general} [DEBUG]")
            logging.error(f"   ❌ Error inesperado en intento {intento}: {error_general}")

            if intento < intentos:
                print(f"   ⏳ Reintentando en {espera_segundos} segundos... [DEBUG]")
                logging.info(f"   ⏳ Reintentando en {espera_segundos} segundos...")
                time.sleep(espera_segundos)
                continue
            else:
                break

    return False, ultimo_error


# ------------------------------------------------------------
# ⚙️ Proceso completo de cargue + mantenimiento
# ------------------------------------------------------------
def run_cargue(database_name: str, archivo_path: str, usuario: str = None):
    """Ejecuta el proceso completo de cargue y mantenimiento."""
    print("🚀🚀🚀 INICIO FUNCIÓN run_cargue - DEBUG LOG 🚀🚀🚀")
    logging.info("🚀🚀🚀 INICIO FUNCIÓN run_cargue - DEBUG LOG 🚀🚀🚀")
    
    start_time = time.time()
    logging.info(f"🚀 Iniciando cargue del archivo: {archivo_path}")
    print(f"🚀 Iniciando cargue del archivo: {archivo_path}")

    # Las fechas se detectarán automáticamente desde la columna "Fecha" del Excel
    logging.info("📅 Las fechas se detectarán automáticamente desde la columna 'Fecha' del Excel")
    print("📅 Las fechas se detectarán automáticamente desde la columna 'Fecha' del Excel")

    cargador = None
    
    try:
        # 🔹 FASE 1: CREAR INSTANCIA DEL CARGADOR
        print("🔧 FASE 1: Creando instancia del cargador... [DEBUG]")
        logging.info("🔧 Fase 1: Creando instancia del cargador...")
        cargador = CargueInfoVentasInsert(
            excel_file=archivo_path,
            database_name=database_name,
            IdtReporteIni=None,  # Se detectará del Excel
            IdtReporteFin=None,  # Se detectará del Excel  
            user_id=usuario or "SYSTEM"
        )
        print("✅ Cargador creado exitosamente [DEBUG]")
        logging.info("✅ Cargador creado exitosamente")

        # 🔹 FASE 1.5: CARGUE PREVIO PARA VALIDACIÓN
        print("🔧 FASE 1.5: Ejecutando cargue previo para validación... [DEBUG]")
        logging.info("🔧 Fase 1.5: Cargando datos en staging para validación...")
        resultado = cargador.procesar_cargue()
        print("✅ Datos cargados en staging para validación [DEBUG]")
        logging.info("✅ Datos cargados correctamente en staging.")
        
        # Obtener fechas desde la tabla staging (detectadas automáticamente del Excel)        
        # Consultar fechas min/max desde la tabla infoventas (staging)
        fecha_ini, fecha_fin = None, None

        def _normalizar_fecha(valor):
            if not valor:
                return None
            if isinstance(valor, datetime):
                return valor.date()
            if hasattr(valor, "to_pydatetime"):
                try:
                    return valor.to_pydatetime().date()
                except Exception:
                    pass
            if hasattr(valor, "year") and hasattr(valor, "month") and hasattr(valor, "day"):
                return datetime(valor.year, valor.month, valor.day).date()
            if isinstance(valor, str):
                texto = valor.strip()
                if not texto or texto.startswith("0000"):
                    return None
                for formato in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
                    try:
                        return datetime.strptime(texto, formato).date()
                    except ValueError:
                        continue
                logging.debug(f"Valor de fecha no interpretable: {valor}")
                return None
            return None

        fecha_min_resultado = _normalizar_fecha(resultado.get('fecha_min')) if isinstance(resultado, dict) else None
        fecha_max_resultado = _normalizar_fecha(resultado.get('fecha_max')) if isinstance(resultado, dict) else None

        if fecha_min_resultado and fecha_max_resultado:
            fecha_ini, fecha_fin = fecha_min_resultado, fecha_max_resultado
        else:
            try:
                with cargador.engine_mysql_bi.connect() as conn:
                    query = text("SELECT MIN(`Fecha`) as fecha_ini, MAX(`Fecha`) as fecha_fin FROM infoventas")
                    result = conn.execute(query).fetchone()

                    fecha_ini_raw = result.fecha_ini
                    fecha_fin_raw = result.fecha_fin
                    fecha_ini_norm = _normalizar_fecha(fecha_ini_raw)
                    fecha_fin_norm = _normalizar_fecha(fecha_fin_raw)

                    if not fecha_ini_norm or not fecha_fin_norm:
                        logging.warning("⚠️ Rango de fechas detectado contiene valores inválidos; la validación recalculará el rango automáticamente")

                    fecha_ini = fecha_ini_norm or fecha_ini_raw
                    fecha_fin = fecha_fin_norm or fecha_fin_raw
            except Exception as e:
                logging.warning(f"No se pudieron obtener fechas de staging: {e}")
                fecha_ini, fecha_fin = None, None
        
        # Registrar estadísticas detalladas
        registros_procesados = resultado.get('registros_procesados', 0)
        registros_insertados = resultado.get('registros_insertados', 0)
        registros_actualizados = resultado.get('registros_actualizados', 0)
        registros_preservados = resultado.get('registros_preservados', 0)
        
        logging.info(f"📊 Registros procesados: {registros_procesados:,}")
        logging.info(f"📊 Registros insertados: {registros_insertados:,}")
        logging.info(f"📊 Registros actualizados: {registros_actualizados:,}")
        logging.info(f"📊 Registros preservados: {registros_preservados:,}")
        
        # Mostrar rango de fechas detectadas
        if fecha_ini and fecha_fin:
            logging.info(f"📅 RANGO DE FECHAS DETECTADAS: {fecha_ini} → {fecha_fin}")
            print(f"📅 RANGO DE FECHAS DETECTADAS: {fecha_ini} → {fecha_fin}")
        else:
            logging.warning("⚠️ No se detectaron fechas válidas del Excel")
        
        # 🔹 FASE 2: EJECUTAR MANTENIMIENTO POST-CARGUE (DIRECTAMENTE, SIN VALIDACIÓN)
        print("🔧 FASE 2: Iniciando mantenimiento post-cargue... [DEBUG]")
        logging.info("🔧 Fase 2: Iniciando mantenimiento post-cargue...")
        ejecutar_mantenimiento_completo(cargador)
        
        # 🔹 FASE 3: DIAGNÓSTICO DE LA VISTA
        print("🔧 FASE 3: Ejecutando diagnóstico de la vista... [DEBUG]")
        logging.info("🔧 Fase 3: Ejecutando diagnóstico de la vista...")
        diagnosticar_vista_infoventas(cargador)
        
        # 🔹 FASE 4: CAPTURAR ESTADÍSTICAS FINALES
        print("🔧 FASE 4: Capturando estadísticas finales... [DEBUG]")
        logging.info("🔧 Fase 4: Capturando estadísticas finales...")
        
        # Calcular tiempo transcurrido
        elapsed_time = time.time() - start_time
        
        # Importar el reporter de email
        from scripts.email_reporter import obtener_estadisticas_tablas
        
        estadisticas_tablas = obtener_estadisticas_tablas(cargador)
        registros_fact = estadisticas_tablas.get('registros_fact', 0)
        registros_dev = estadisticas_tablas.get('registros_dev', 0)
        registros_staging = estadisticas_tablas.get('registros_staging', 0)
        detalles_tablas = estadisticas_tablas.get('detalles_tablas', [])
        
        # Logging detallado de estadísticas
        logging.info("=" * 80)
        logging.info("📊 === ESTADÍSTICAS FINALES DE CARGUE ===")
        logging.info("=" * 80)
        logging.info(f"📅 Período procesado: {fecha_ini} → {fecha_fin}")
        logging.info(f"⏱️  Duración total: {elapsed_time:.2f} segundos")
        logging.info("")
        logging.info("📝 RESUMEN DE INSERCIÓN:")
        logging.info(f"   • Registros procesados: {registros_procesados:,}")
        logging.info(f"   • Registros insertados: {registros_insertados:,}")
        logging.info(f"   • Registros actualizados: {registros_actualizados:,}")
        logging.info(f"   • Registros preservados: {registros_preservados:,}")
        logging.info("")
        logging.info("📦 DISTRIBUCIÓN POR TABLA CLASIFICADA:")
        logging.info(f"   • Registros en _fact: {registros_fact:,}")
        logging.info(f"   • Registros en _dev: {registros_dev:,}")
        logging.info(f"   • Total clasificado: {registros_fact + registros_dev:,}")
        logging.info(f"   • Registros en staging (post-limpieza): {registros_staging:,}")
        logging.info("")
        logging.info("📋 DETALLES POR TABLA:")
        for tabla_info in detalles_tablas:
            tabla_nombre = tabla_info.get('tabla', '')
            tipo = tabla_info.get('tipo', '')
            registros = tabla_info.get('registros', 0)
            logging.info(f"   • {tabla_nombre}: {registros:,} registros [{tipo}]")
        logging.info("=" * 80)
        
        # 🔹 FASE 5: REPORTE FINAL CON ESTADÍSTICAS
        print(f"🎉 PROCESO COMPLETADO EXITOSAMENTE en {elapsed_time:.2f} segundos [DEBUG]")
        logging.info(f"🎉 PROCESO COMPLETADO EXITOSAMENTE en {elapsed_time:.2f} segundos")
        
    except Exception as e:
        elapsed_time = time.time() - start_time
        print(f"❌ ERROR CRÍTICO en el proceso principal: {e} [DEBUG]")
        logging.error(f"❌ ERROR CRÍTICO en el proceso principal: {e}", exc_info=True)
        logging.error(f"❌ Tiempo hasta error: {elapsed_time:.2f} segundos")
        raise e
    finally:
        # Limpieza final
        print("🧹 Ejecutando limpieza final... [DEBUG]")
        if cargador and hasattr(cargador, 'engine_mysql_bi'):
            try:
                cargador.engine_mysql_bi.dispose()
                logging.info("🔒 Engine de base de datos cerrado correctamente.")
            except Exception:
                pass
        print("🏁 FIN FUNCIÓN run_cargue [DEBUG]")


def ejecutar_mantenimiento_completo(cargador):
    """Ejecuta el mantenimiento completo post-cargue con múltiples métodos de respaldo."""
    print("🧹 === INICIANDO FUNCIÓN ejecutar_mantenimiento_completo [DEBUG] ===")
    logging.info("🧹 === INICIANDO MANTENIMIENTO POST-CARGUE ===")
    
    # Verificar registros antes del mantenimiento
    registros_antes = 0
    try:
        print("📊 Verificando registros antes del mantenimiento... [DEBUG]")
        conn = cargador.engine_mysql_bi.raw_connection()
        try:
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT COUNT(*) FROM infoventas;")
                registros_antes = cursor.fetchone()[0]
                print(f"📊 Registros en infoventas ANTES: {registros_antes} [DEBUG]")
                logging.info(f"📊 Registros en infoventas ANTES del mantenimiento: {registros_antes}")
            finally:
                cursor.close()
        finally:
            conn.close()
    except Exception as e:
        print(f"❌ Error verificando registros antes: {e} [DEBUG]")
        logging.error(f"❌ Error verificando registros antes del mantenimiento: {e}")
        registros_antes = -1

    # Método 1: Intentar con conexión cruda
    mantenimiento_exitoso = False
    
    try:
        print("🔧 Método 1: Ejecutando con raw_connection y reintentos... [DEBUG]")
        logging.info("🔧 Método 1: Ejecutando con raw_connection y reintentos...")
        exito_raw, error_raw = ejecutar_procedimiento_con_reintentos(
            cargador,
            "CALL sp_infoventas_full_maintenance();",
            intentos=3,
            espera_segundos=60,
        )

        if not exito_raw:
            if error_raw:
                raise error_raw
            raise RuntimeError("No se pudo ejecutar el procedimiento, sin detalle adicional")

        mantenimiento_exitoso = True

    except Exception as e:
        print(f"❌ Error en Método 1: {e} [DEBUG]")
        logging.error(f"❌ Error en Método 1 (raw_connection): {e}")
        logging.error(f"❌ Tipo de error: {type(e).__name__}")
        
        # Método 2: Intentar con SQLAlchemy text()
        try:
            print("🔧 Método 2: Intentando con SQLAlchemy text()... [DEBUG]")
            logging.info("🔧 Método 2: Intentando con SQLAlchemy text()...")
            with cargador.engine_mysql_bi.begin() as connection:
                for comando in (
                    "SET SESSION wait_timeout = 7200",
                    "SET SESSION interactive_timeout = 7200",
                    "SET SESSION net_read_timeout = 600",
                    "SET SESSION net_write_timeout = 600",
                    "SET SESSION innodb_lock_wait_timeout = 900",
                ):
                    connection.exec_driver_sql(comando)

                print("📡 Ejecutando: CALL sp_infoventas_full_maintenance() con text() [DEBUG]")
                logging.info("📡 Ejecutando: CALL sp_infoventas_full_maintenance() con text()")
                result = connection.execute(text("CALL sp_infoventas_full_maintenance();"))
                print("✅ Procedimiento ejecutado con SQLAlchemy text() [DEBUG]")
                logging.info("✅ Procedimiento ejecutado con SQLAlchemy text()")
                
                # Verificar si hay resultados
                try:
                    cursor = getattr(result, "cursor", None)
                    if cursor is not None:
                        while True:
                            try:
                                rows = cursor.fetchall()
                                if rows:
                                    print(f"📋 Resultados: {rows} [DEBUG]")
                                    logging.info(f"📋 Resultados: {rows}")
                            except Exception:
                                pass

                            if not cursor.nextset():
                                break
                    else:
                        rows = result.fetchall()
                        if rows:
                            print(f"📋 Resultados: {rows} [DEBUG]")
                            logging.info(f"📋 Resultados: {rows}")
                except Exception as warn_err:
                    print(f"📋 Sin resultados específicos [DEBUG] ({warn_err})")
                    logging.info(f"📋 Sin resultados específicos ({warn_err})")
                finally:
                    try:
                        result.close()
                    except Exception:
                        pass
                
                mantenimiento_exitoso = True
                        
        except Exception as e2:
            print(f"❌ Error en Método 2: {e2} [DEBUG]")
            logging.error(f"❌ Error en Método 2 (SQLAlchemy text): {e2}")
            logging.error(f"❌ Tipo de error: {type(e2).__name__}")
            
            # Método 3: Verificar que el procedimiento existe y diagnosticar
            try:
                print("🔧 Método 3: Diagnóstico del procedimiento... [DEBUG]")
                logging.info("🔧 Método 3: Diagnóstico del procedimiento...")
                conn = cargador.engine_mysql_bi.raw_connection()
                try:
                    cursor = conn.cursor()
                    try:
                        cursor.execute("SHOW PROCEDURE STATUS WHERE Name = 'sp_infoventas_full_maintenance';")
                        proc_info = cursor.fetchall()
                        if proc_info:
                            print(f"✅ Procedimiento encontrado: {proc_info} [DEBUG]")
                            logging.info(f"✅ Procedimiento encontrado: {proc_info}")
                        else:
                            print("❌ Procedimiento sp_infoventas_full_maintenance NO existe [DEBUG]")
                            logging.error("❌ Procedimiento sp_infoventas_full_maintenance NO existe")
                            
                        # Intentar procedimiento simple para probar conectividad
                        cursor.execute("SELECT 'TEST_CONNECTION' as test;")
                        test_result = cursor.fetchone()
                        print(f"✅ Test de conexión exitoso: {test_result} [DEBUG]")
                        logging.info(f"✅ Test de conexión exitoso: {test_result}")
                    finally:
                        cursor.close()
                finally:
                    conn.close()
                        
            except Exception as e3:
                print(f"❌ Error en Método 3: {e3} [DEBUG]")
                logging.error(f"❌ Error en Método 3 (verificación): {e3}")

    # Verificar resultado final del mantenimiento
    try:
        print("📊 Verificando resultado final... [DEBUG]")
        conn = cargador.engine_mysql_bi.raw_connection()
        try:
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT COUNT(*) FROM infoventas;")
                registros_despues = cursor.fetchone()[0]
                print(f"📊 Registros en infoventas DESPUÉS: {registros_despues} [DEBUG]")
                logging.info(f"📊 Registros en infoventas DESPUÉS del mantenimiento: {registros_despues}")
            finally:
                cursor.close()
        finally:
            conn.close()

        if registros_despues == 0:
            print("✅ Mantenimiento completado. Tabla infoventas limpia. [DEBUG]")
            logging.info("✅ Mantenimiento completado. Tabla infoventas limpia.")
            mantenimiento_exitoso = True
        elif registros_antes > 0 and registros_despues < registros_antes:
            print(f"✅ Mantenimiento parcial. Reducidos de {registros_antes} a {registros_despues} registros. [DEBUG]")
            logging.info(f"✅ Mantenimiento parcial. Reducidos de {registros_antes} a {registros_despues} registros.")
            mantenimiento_exitoso = True
        elif registros_antes > 0 and registros_despues == registros_antes:
            print(f"⚠️ Mantenimiento posiblemente no ejecutado. Registros sin cambios: {registros_despues} [DEBUG]")
            logging.warning(f"⚠️ Mantenimiento posiblemente no ejecutado. Registros sin cambios: {registros_despues}")
        else:
            print(f"📊 Estado post-mantenimiento: {registros_despues} registros [DEBUG]")
            logging.info(f"📊 Estado post-mantenimiento: {registros_despues} registros")
            
    except Exception as e:
        print(f"❌ Error verificando resultado final: {e} [DEBUG]")
        logging.error(f"❌ Error verificando resultado final: {e}")
    
    if mantenimiento_exitoso:
        print("🎉 === MANTENIMIENTO COMPLETADO EXITOSAMENTE === [DEBUG]")
        logging.info("🎉 === MANTENIMIENTO COMPLETADO EXITOSAMENTE ===")
    else:
        print("⚠️ === MANTENIMIENTO CON ERRORES - REVISAR LOGS === [DEBUG]")
        logging.warning("⚠️ === MANTENIMIENTO CON ERRORES - REVISAR LOGS ===")
    
    print("🏁 FIN FUNCIÓN ejecutar_mantenimiento_completo [DEBUG]")
    return mantenimiento_exitoso


# ============================================================
# 🔍 DIAGNÓSTICO: Verificar composición de la vista
# ============================================================
def diagnosticar_vista_infoventas(cargador):
    """
    Verifica que vw_infoventas SOLO contenga tablas _fact y _dev.
    Detecta si hay tablas anuales (infoventas_YYYY) incluidas incorrectamente.
    """
    print("\n" + "="*70)
    print(f"{TerminalColors.BOLD}🔍 DIAGNÓSTICO: Composición de vw_infoventas{TerminalColors.ENDC}")
    print("="*70)
    
    try:
        conn = cargador.engine_mysql_bi.raw_connection()
        try:
            cursor = conn.cursor()
            try:
                # 1️⃣ Obtener definición de la vista
                print(f"\n{TerminalColors.OKBLUE}1️⃣ Estructura de vw_infoventas:{TerminalColors.ENDC}")
                cursor.execute("SHOW CREATE VIEW vw_infoventas;")
                vista_def = cursor.fetchone()
                if vista_def:
                    vista_sql = vista_def[1]
                    # Contar UNION ALL para determinar tablas incluidas
                    num_uniones = vista_sql.count(" UNION ALL ")
                    num_tablas = num_uniones + 1
                    print(f"   📊 Tablas en la vista: {num_tablas}")
                    
                    # Verificar si incluye tablas anuales (❌ MAL)
                    if "FROM `infoventas_2024`" in vista_sql or \
                       "FROM `infoventas_2025`" in vista_sql or \
                       "FROM `infoventas_2026`" in vista_sql:
                        print(f"{TerminalColors.FAIL}   ❌ ERROR: La vista incluye tablas anuales completas (infoventas_YYYY){TerminalColors.ENDC}")
                        print(f"{TerminalColors.FAIL}   Esto causa DUPLICACIÓN de datos.{TerminalColors.ENDC}")
                    else:
                        print(f"{TerminalColors.OKGREEN}   ✅ La vista NO incluye tablas anuales completas.{TerminalColors.ENDC}")
                    
                    # Verificar si incluye _fact y _dev
                    fact_count = vista_sql.count("_fact")
                    dev_count = vista_sql.count("_dev")
                    print(f"   📊 Tablas _fact: {fact_count}")
                    print(f"   📊 Tablas _dev: {dev_count}")
                    
                    if fact_count > 0 and dev_count > 0:
                        print(f"{TerminalColors.OKGREEN}   ✅ La vista incluye correctamente tablas _fact y _dev.{TerminalColors.ENDC}")
                    else:
                        print(f"{TerminalColors.WARNING}   ⚠️ La vista podría no tener tablas _fact o _dev.{TerminalColors.ENDC}")

                # 2️⃣ Listar tablas que se detectaron
                print(f"\n{TerminalColors.OKBLUE}2️⃣ Tablas detectadas en la base de datos:{TerminalColors.ENDC}")
                cursor.execute("""
                    SELECT table_name FROM information_schema.tables
                    WHERE table_schema = DATABASE() AND table_name LIKE 'infoventas\\_%' ESCAPE '\\\\'
                    ORDER BY table_name;
                """)
                tablas = cursor.fetchall()
                
                tablas_anuales = []
                tablas_fact_dev = []
                
                for (tabla,) in tablas:
                    if tabla.endswith("_fact") or tabla.endswith("_dev"):
                        tablas_fact_dev.append(tabla)
                    else:
                        tablas_anuales.append(tabla)
                
                print(f"\n   📋 Tablas anuales (fuente, NO en vista): {len(tablas_anuales)}")
                for tbl in sorted(tablas_anuales):
                    print(f"      • {tbl}")
                
                print(f"\n   📋 Tablas _fact/_dev (en vista): {len(tablas_fact_dev)}")
                for tbl in sorted(tablas_fact_dev):
                    print(f"      • {tbl}")

                # 3️⃣ Contar registros en cada tabla
                print(f"\n{TerminalColors.OKBLUE}3️⃣ Conteo de registros:{TerminalColors.ENDC}")
                
                # Vista
                cursor.execute("SELECT COUNT(*) FROM vw_infoventas;")
                vista_count = cursor.fetchone()[0]
                print(f"   📊 vw_infoventas: {vista_count:,} registros")
                
                # Staging
                cursor.execute("SELECT COUNT(*) FROM infoventas;")
                staging_count = cursor.fetchone()[0]
                print(f"   📊 infoventas (staging): {staging_count:,} registros")
                
                # Totales por tipo
                total_fact = 0
                total_dev = 0
                for tbl in sorted(tablas_fact_dev):
                    cursor.execute(f"SELECT COUNT(*) FROM `{tbl}`;")
                    count = cursor.fetchone()[0]
                    if "_fact" in tbl:
                        total_fact += count
                    elif "_dev" in tbl:
                        total_dev += count
                
                print(f"   📊 Total _fact: {total_fact:,} registros")
                print(f"   📊 Total _dev: {total_dev:,} registros")
                print(f"   📊 Total en vista: {vista_count:,} registros (debe = fact + dev)")
                
                # Validación
                if vista_count == (total_fact + total_dev):
                    print(f"{TerminalColors.OKGREEN}   ✅ Consistencia verificada.{TerminalColors.ENDC}")
                else:
                    print(f"{TerminalColors.WARNING}   ⚠️ Posible inconsistencia: vista={vista_count}, suma fact+dev={total_fact + total_dev}{TerminalColors.ENDC}")

            finally:
                cursor.close()
        finally:
            conn.close()
            
    except Exception as e:
        print(f"{TerminalColors.FAIL}❌ Error en diagnóstico: {e}{TerminalColors.ENDC}")
        logging.error(f"❌ Error en diagnóstico de vista: {e}", exc_info=True)
    
    print("="*70 + "\n")


# ============================================================
# 🧩 Lógica principal con CLI
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Carga automatizada de InfoVentas")
    parser.add_argument("--base", required=True, help="Nombre de la base de datos (ej: bi_distrijass)")
    parser.add_argument("--archivo", help="Ruta de un archivo específico")

    parser.add_argument("--carpeta", help="Ruta de carpeta con múltiples archivos")
    parser.add_argument("--usuario", help="Usuario que ejecuta el proceso (por defecto SYSTEM)")

    args = parser.parse_args()
    database_name = args.base
    usuario = args.usuario or "SYSTEM"

    if args.archivo:
        run_cargue(database_name, args.archivo, usuario)
    elif args.carpeta:
        archivos = sorted([
            os.path.join(args.carpeta, f)
            for f in os.listdir(args.carpeta)
            if f.endswith(".xlsx") or f.endswith(".csv")
        ])
        for archivo in archivos:
            run_cargue(database_name, archivo, usuario)
            time.sleep(3)  # Pequeña pausa entre archivos
    else:
        logging.error("❌ Debes indicar --archivo o --carpeta.")


if __name__ == "__main__":
    main()
